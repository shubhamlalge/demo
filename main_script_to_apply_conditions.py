import pandas as pd
import openpyxl
from sqlalchemy import create_engine, text
import datetime
import re

# Database connection URI
db_uri = "mysql+mysqlconnector://root:root@localhost:3307/yoandb"

# Create a SQLAlchemy engine with connection pooling
engine = create_engine(db_uri)
connection = engine.connect()


def load_conditions_from_excel(file_path, column_number):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    conditions = []
    for row_idx, row in enumerate(sheet.iter_rows(min_col=column_number, max_col=column_number, values_only=True),
                                  start=1):
        if row_idx == 1:  # Skip the first row (header)
            continue
        for cell_value in row:
            if cell_value is not None:
                conditions.append(cell_value)
            else:
                break
    return conditions


# Function to process data in batches
def fetch_all_data(engine, table):
    try:
        with engine.connect() as conn:
            # Fetch all data from the table
            query = f"SELECT * FROM {table}"
            result = conn.execute(text(query))
            data = result.fetchall()
        return data
    except Exception as e:
        print(f"Error fetching data from {table}: {e}")
        return []


def main():
    try:
        file_path = r'\\yoandc\Campaigns\#Yoanone\dipesh\JT_For_Extraction.xlsm'
        # Load input conditions from the first sheet of Excel
        workbook = openpyxl.load_workbook(
            file_path)
        print("Execution Start")
        start_time = datetime.datetime.now().strftime("%M%S")
        start_time_int = int(start_time)
        print("Start time:", start_time)

        # Load input conditions from Excel
        first_conditions = load_conditions_from_excel(
            file_path, column_number=1)
        second_conditions = load_conditions_from_excel(
            file_path, column_number=2)
        third_conditions = load_conditions_from_excel(
            file_path, column_number=3)
        fourth_conditions = load_conditions_from_excel(
            file_path, column_number=4)
        fifth_conditions = load_conditions_from_excel(
            file_path, column_number=5)
        sixth_conditions = load_conditions_from_excel(
            file_path, column_number=6)
        seven_conditions = load_conditions_from_excel(
            file_path, column_number=7)
        email_conditions = load_conditions_from_excel(
            file_path, column_number=8)
        jt_link_conditions = load_conditions_from_excel(
            file_path, column_number=9)
        fl_domain_conditions = load_conditions_from_excel(
            file_path, column_number=10)
        fl_company_conditions = load_conditions_from_excel(
            file_path, column_number=11)

        # Load 'Job Level' sheet from Excel
        job_level_sheet = workbook['Job Level']
        mapped_conditions = []
        for condition in first_conditions:
            for row in job_level_sheet.iter_rows(values_only=True):
                if condition.lower() in str(row[0]).lower():
                    mapped_conditions.append(row[1])
        print(mapped_conditions)

        company_size = workbook['companySize']
        mapped_fourth_conditions = []
        for condition in fourth_conditions:
            for row in company_size.iter_rows(values_only=True):
                if condition.lower() in str(row[0]).lower():
                    mapped_fourth_conditions.append(row[1])

        second_sheet = workbook['JT']  # Provide the actual name of your second sheet
        matched_column_indexes = []
        matched_conditions = []
        # Iterate through rows in the second sheet
        for row_idx, row in enumerate(second_sheet.iter_rows(min_row=1, max_row=1, values_only=True), start=1):
            # Iterate through cells in the row
            for idx, cell_value in enumerate(row, start=1):
                # Check if the cell value matches any of the second_conditions
                if cell_value in second_conditions:
                    matched_column_indexes.append(idx)

        # Iterate through matched columns
        for matched_column_index in matched_column_indexes:

            # Get all conditions from the current matched column
            for row_idx, cell_value in enumerate(
                    second_sheet.iter_rows(min_col=matched_column_index, max_col=matched_column_index,
                                           values_only=True),
                    start=1):
                if row_idx == 1:  # Skip the first row (header)
                    continue
                if cell_value[0] is not None:
                    matched_conditions.append(cell_value[0])
                else:
                    break

        all_results_df = pd.DataFrame()  # Create an empty DataFrame

        count = 1
        for i in range(1, 22):
            # Process data in batches
            table = f"yoan_one_table_{count}"
            print(table)
            all_result = fetch_all_data(engine, table)

            count += 1
            # Create a DataFrame from the results
            header = ['Date', 'Salutation', 'First Name', 'Last Name', 'Email', 'Company Name', 'Address_1',
                      'City', 'State', 'Zip Code', 'COUNTRY', 'Industry', 'Standard_Industry',
                      'Job_Title', 'Job Title Level', 'Job Title Department', 'Employee Size', 'Revenue_Size',
                      'Phone NO', 'Direct_Dial_Extension', 'SIC_Code', 'NAICS_Code', 'Job Title Link',
                      'Employee Size Link',
                      'Revenue_Size_Link', 'VV Status', 'Final Status', 'id', 'domain', 'FirstLastDomain',
                      'FirstLastCompany']

            results_df = pd.DataFrame(all_result, columns=header)
            # Concatenate the new results with the existing DataFrame
            all_results_df = pd.concat([all_results_df, results_df], ignore_index=True)

        try:
            print("Data filtration start")

            # Function to apply regex pattern using re.findall
            def apply_regex(column, pattern):
                return column.apply(lambda x: bool(re.findall(pattern, str(x), flags=re.IGNORECASE)))

            try:
                if sixth_conditions:
                    all_results_df = all_results_df[all_results_df['domain'].isin(sixth_conditions)]
                    print("six", all_results_df.shape)
            except Exception as e:
                print("Exception in Tal", e)

            try:
                if third_conditions:
                    all_results_df = all_results_df[all_results_df['COUNTRY'].isin(third_conditions)]
                    print("third", all_results_df.shape)
            except Exception as e:
                print("Exception in country", e)

            try:
                if mapped_fourth_conditions:
                    all_results_df = all_results_df[all_results_df['Employee Size'].isin(mapped_fourth_conditions)]
            except Exception as e:
                print("Exception in employee size", e)

            try:
                # Check if seventh_conditions is not empty, apply it to the DataFrame
                if seven_conditions:
                    all_results_df = all_results_df[~all_results_df['domain'].isin(seven_conditions)]
            except Exception as e:
                print("Exception in suppression", e)

            try:
                # Apply the fifth condition using str.contains
                if mapped_conditions:
                    condition_series = []
                    for condition in mapped_conditions:
                        pattern = f".*{condition}.*"
                        condition_series.append(
                            all_results_df['Job_Title'].str.contains(pattern, case=False, na=False, regex=True))
                    if condition_series:
                        # Combine conditions using logical OR
                        final_condition = pd.DataFrame(condition_series).any(axis=0)
                        all_results_df = all_results_df[final_condition]
            except Exception as e:
                print("Error in industry condition:", e)

            try:
                # Apply the fifth condition using re.findall
                if fifth_conditions:
                    for condition in fifth_conditions:
                        pattern = f"{condition}.*"
                        all_results_df = all_results_df[~apply_regex(all_results_df['Industry'], pattern)]
            except Exception as e:
                print("Exception in Industry", e)

            print("After 1st condition:", all_results_df.shape)

            try:
                # Apply the fifth condition using str.contains
                if matched_conditions:
                    condition_series = []
                    for condition in matched_conditions:
                        pattern = f".*{condition}.*"
                        condition_series.append(
                            all_results_df['Job_Title'].str.contains(pattern, case=False, na=False, regex=True))
                    if condition_series:
                        # Combine conditions using logical OR
                        final_condition = pd.DataFrame(condition_series).any(axis=0)
                        all_results_df = all_results_df[final_condition]
            except Exception as e:
                print("Error in industry condition:", e)
            try:
                # Check if email_conditions is not empty, apply it to the DataFrame
                if email_conditions:
                    all_results_df['Email'] = all_results_df['Email'].str.lower()
                    all_results_df = all_results_df[~all_results_df['Email'].isin(email_conditions)].drop_duplicates(
                        'Email')
            except Exception as e:
                print("Exception in email suppression", e)

            try:
                # Check if jt_link_conditions is not empty, apply it to the DataFrame
                if jt_link_conditions:
                    all_results_df = all_results_df[~all_results_df['Job Title Link'].isin(jt_link_conditions)]
            except Exception as e:
                print("Exception in JT link suppression", e)

            try:
                # Check if fl_domain_conditions is not empty, apply it to the DataFrame
                if fl_domain_conditions:
                    all_results_df = all_results_df[
                        ~all_results_df['FirstLastDomain'].isin(fl_domain_conditions)].drop_duplicates(
                        'FirstLastDomain')
            except Exception as e:
                print("Exception in FL_domain suppression", e)

            try:
                # Check if fl_company_conditions is not empty, apply it to the DataFrame
                if fl_company_conditions:
                    all_results_df = all_results_df[
                        ~all_results_df['FirstLastCompany'].isin(fl_company_conditions)].drop_duplicates(
                        'FirstLastCompany')
            except Exception as e:
                print("Exception in FL_Company suppression", e)

            all_results_df = all_results_df.drop_duplicates('Job Title Link')
        except Exception as e:
            print(e)

        output_file_name = input("Enter file name:")
        output_file_path = fr'\\yoandc\Campaigns\#Yoanone\dipesh\{output_file_name}'

        all_results_df.to_excel(output_file_path, index=False)

        end_time = datetime.datetime.now().strftime("%M%S")
        end_time_int = int(end_time)
        print("end time:", end_time)
        total_time_script_takes = abs(start_time_int - end_time_int)
        print("Total time takes:", total_time_script_takes)
        # Print message
        print(f"Data saved into Excel file: {output_file_path}")

    except Exception as e:
        print(e)


if __name__ == '__main__':
    main()
